# ruff: noqa: E501
# mypy: disable-error-code="no-untyped-def,no-untyped-call,import-untyped,no-any-return,operator,attr-defined"
from __future__ import annotations

import csv
import io
import json
from datetime import UTC, date, datetime
from decimal import Decimal
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy import delete, select

from app.core.errors import ApplicationError
from app.models.imports import ImportJob, ImportStatus
from app.models.registers import (
    Expense,
    ExpenseDocumentType,
    Offer,
    PaymentStatus,
    Procedure,
    RecordSource,
    RecordStatus,
    Sale,
    Supplier,
)

FIELD_MAPS = {
    "offers": {
        "name": ["name", "nom", "offre", "article", "produit", "designation", "désignation"],
        "price": ["price", "prix", "tarif", "prix_unitaire", "pu", "prix vente"],
        "currency": ["currency", "devise", "monnaie"],
        "category": ["category", "categorie", "catégorie", "famille", "rayon"],
        "stock_quantity": ["stock_quantity", "stock", "quantite", "quantité", "qte"],
        "cost_price": ["cost_price", "prix_achat", "cout", "coût"],
    },
    "sales": {
        "reference": ["reference", "référence", "ref", "n_facture", "num_facture", "facture", "recu", "reçu"],
        "sale_date": ["sale_date", "date", "date_vente", "date vente", "date_facture"],
        "client_name": ["client_name", "client", "nom_client", "acheteur", "nom"],
        "item_label": ["item_label", "produit", "service", "offre", "article", "designation", "désignation", "libelle"],
        "quantity": ["quantity", "quantite", "quantité", "qte", "nombre"],
        "unit_price": ["unit_price", "prix_unitaire", "pu", "prix", "tarif"],
        "discount": ["discount", "remise", "rabais", "reduction"],
        "payment_method": ["payment_method", "mode_paiement", "mode paiement", "reglement", "règlement"],
        "payment_status": ["payment_status", "etat_paiement", "statut", "statut_paiement", "paye", "payé"],
    },
    "depenses": {
        "reference": ["reference", "référence", "ref", "n_piece", "piece", "justificatif"],
        "expense_date": ["expense_date", "date", "date_depense", "date dépense"],
        "beneficiary": ["beneficiary", "beneficiaire", "bénéficiaire", "fournisseur", "prestataire", "nom"],
        "category": ["category", "categorie", "catégorie", "type_charge", "type"],
        "amount": ["amount", "montant", "total", "prix", "ttc"],
        "payment_method": ["payment_method", "mode_paiement", "mode paiement", "reglement"],
        "payment_status": ["payment_status", "statut", "etat_paiement"],
    },
    "fournisseurs": {
        "name": ["name", "nom", "fournisseur", "raison_sociale", "entreprise"],
        "category": ["category", "categorie", "catégorie", "secteur", "activite"],
        "contact_name": ["contact_name", "contact", "interlocuteur", "responsable"],
        "phone": ["phone", "telephone", "téléphone", "tel", "mobile", "whatsapp"],
        "email": ["email", "e-mail", "courriel", "mail"],
        "address": ["address", "adresse", "ville", "localisation"],
    },
    "procedures": {
        "title": ["title", "titre", "procedure", "procédure", "processus", "nom"],
        "objective": ["objective", "objectif", "but", "finalite"],
        "department": ["department", "service", "departement", "département", "pole"],
        "responsible_user_id": ["responsible_user_id", "responsable", "porteur"],
    },
}

# Aliases for registers
FIELD_MAPS["expenses"] = FIELD_MAPS["depenses"]
FIELD_MAPS["products"] = FIELD_MAPS["offers"]
FIELD_MAPS["suppliers"] = FIELD_MAPS["fournisseurs"]

REQUIRED = {
    "offers": {"name"},
    "products": {"name"},
    "sales": {"item_label"},
    "depenses": {"beneficiary", "amount"},
    "expenses": {"beneficiary", "amount"},
    "fournisseurs": {"name"},
    "suppliers": {"name"},
    "procedures": {"title"},
}


class ImportService:
    def parse(self, filename: str, content: bytes) -> list[dict[str, object]]:
        extension = filename.lower().rsplit(".", 1)[-1]
        
        # CSV & TSV (Multi-delimiter detection)
        if extension in {"csv", "tsv", "txt"}:
            text = content.decode("utf-8-sig", errors="replace")
            # Auto-detect delimiter
            sample = text[:2048]
            delimiter = ","
            if "\t" in sample and sample.count("\t") > sample.count(","):
                delimiter = "\t"
            elif ";" in sample and sample.count(";") > sample.count(","):
                delimiter = ";"
            elif "|" in sample and sample.count("|") > sample.count(","):
                delimiter = "|"
            
            reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
            rows = [dict(row) for row in reader if any(v and str(v).strip() for v in row.values())]
            return rows

        # JSON
        if extension == "json":
            payload = json.loads(content.decode("utf-8-sig", errors="replace"))
            if not isinstance(payload, list) or not all(isinstance(row, dict) for row in payload):
                raise ApplicationError(
                    "invalid_json",
                    "Le JSON doit contenir une liste d’objets",
                    422,
                )
            return payload

        # XLSX & XLS
        if extension in {"xlsx", "xls"}:
            try:
                workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                worksheet = workbook.active
                values = list(worksheet.iter_rows(values_only=True))
                if not values:
                    return []
                # Clean headers
                headers = [str(value or f"Col_{idx+1}").strip() for idx, value in enumerate(values[0])]
                rows = []
                for row in values[1:]:
                    if any(v is not None and str(v).strip() for v in row):
                        rows.append({headers[index]: value for index, value in enumerate(row) if index < len(headers)})
                return rows
            except Exception as e:
                raise ApplicationError(
                    "excel_parse_error",
                    f"Erreur de lecture du fichier Excel : {str(e)}",
                    422,
                ) from e

        raise ApplicationError(
            "unsupported_file",
            "Formats acceptés : CSV, TSV, XLSX, XLS, JSON ou TXT structuré.",
            415,
        )

    def suggest(self, headers: list[str], register_type: str) -> dict[str, str]:
        reg_key = "depenses" if register_type == "expenses" else ("offers" if register_type == "products" else register_type)
        reg_map = FIELD_MAPS.get(reg_key, {})
        suggestions: dict[str, str] = {}
        for header in headers:
            normalized = header.strip().lower().replace("_", " ").replace("-", " ")
            for field, aliases in reg_map.items():
                if normalized in aliases or header.strip().lower() in aliases:
                    suggestions[header] = field
                    break
        return suggestions

    async def preview(
        self,
        session,
        organization_id: str,
        user_id: str,
        register_type: str,
        filename: str,
        content: bytes,
    ) -> tuple[ImportJob, list[str], dict[str, str]]:
        reg_key = "depenses" if register_type == "expenses" else ("offers" if register_type == "products" else register_type)
        if reg_key not in FIELD_MAPS:
            raise ApplicationError("invalid_register", f"Registre invalide : {register_type}", 400)
        
        rows = self.parse(filename, content)
        if not rows:
            raise ApplicationError("empty_file", "Le fichier ne contient aucune ligne exploitable", 422)
        
        headers = list(rows[0].keys())
        mapping = self.suggest(headers, reg_key)
        
        duplicates: list[int] = []
        seen: set[tuple[str, ...]] = set()
        for row_number, row in enumerate(rows, start=2):
            signature = tuple(str(row.get(header, "")) for header in headers)
            if signature in seen:
                duplicates.append(row_number)
            seen.add(signature)
            
        job = ImportJob(
            organization_id=organization_id,
            register_type=reg_key,
            filename=filename,
            status=ImportStatus.PREVIEW,
            column_mapping=mapping,
            preview_rows=rows[:500],
            duplicate_rows=duplicates,
            row_count=len(rows),
            created_by_user_id=user_id,
        )
        session.add(job)
        await session.commit()
        await session.refresh(job)
        return job, headers, mapping

    async def confirm(
        self,
        session,
        organization_id: str,
        user_id: str,
        job_id: str,
        mapping: dict[str, str],
    ) -> ImportJob:
        job = await session.scalar(
            select(ImportJob).where(
                ImportJob.id == job_id,
                ImportJob.organization_id == organization_id,
            )
        )
        if job is None:
            raise ApplicationError("import_not_found", "Import introuvable", 404)
        if job.status != ImportStatus.PREVIEW:
            raise ApplicationError("import_not_pending", "Import déjà validé ou archivé", 409)
        
        required_fields = REQUIRED.get(job.register_type, set())
        mapped_targets = set(mapping.values())
        missing = required_fields - mapped_targets
        if missing:
            raise ApplicationError(
                "mapping_incomplete",
                f"Veuillez faire correspondre les champs obligatoires suivants : {', '.join(sorted(missing))}",
                422,
            )
            
        imported_ids: list[str] = []
        errors: list[dict[str, object]] = []
        
        for row_number, row in enumerate(job.preview_rows, start=2):
            data = {target: row.get(source) for source, target in mapping.items() if target}
            try:
                record = self._build(job.register_type, organization_id, user_id, data)
                session.add(record)
                await session.flush()
                imported_ids.append(record.id)
            except Exception as exc:
                errors.append({"row": row_number, "message": str(exc)})
                
        if errors and len(errors) > len(job.preview_rows) * 0.5:
            await session.rollback()
            job = await session.scalar(
                select(ImportJob).where(
                    ImportJob.id == job_id,
                    ImportJob.organization_id == organization_id,
                )
            )
            assert job is not None
            job.status = ImportStatus.FAILED
            job.errors = errors
            job.failure_reason = f"{len(errors)} lignes n'ont pas pu être interprétées"
            await session.commit()
            await session.refresh(job)
            return job
            
        job.status = ImportStatus.COMPLETED
        job.column_mapping = mapping
        job.imported_record_ids = imported_ids
        job.completed_at = datetime.now(UTC)
        await session.commit()
        await session.refresh(job)
        return job

    def _build(
        self,
        register_type: str,
        organization_id: str,
        user_id: str,
        data: dict[str, object],
    ):
        if register_type in {"offers", "products"}:
            qty = self._decimal(data.get("stock_quantity"), default=Decimal("0.00"))
            return Offer(
                organization_id=organization_id,
                name=self._required_text(data, "name"),
                price=self._decimal(data.get("price"), nullable=True),
                currency=str(data.get("currency") or "XOF").upper(),
                category=self._optional_text(data.get("category")),
                track_stock=qty > 0,
                stock_quantity=qty,
                cost_price=self._decimal(data.get("cost_price"), nullable=True),
                source=RecordSource.EXCEL,
                status=RecordStatus.VALIDATED,
                created_by_user_id=user_id,
                updated_by_user_id=user_id,
            )
            
        if register_type == "sales":
            quantity = self._decimal(data.get("quantity"), default=Decimal("1.00"))
            unit_price = self._decimal(data.get("unit_price"), default=Decimal("0.00"))
            discount = self._decimal(data.get("discount"), default=Decimal("0.00"))
            total_amt = max(Decimal("0.00"), quantity * unit_price - discount)
            payment_status = self._payment_status(data.get("payment_status"))
            paid_amount = total_amt if payment_status == PaymentStatus.PAID else Decimal("0.00")
            ref = str(data.get("reference") or f"IMP-VENTE-{uuid4().hex[:6].upper()}").strip()
            
            return Sale(
                organization_id=organization_id,
                reference=ref,
                sale_date=self._date(data.get("sale_date")),
                client_name=self._optional_text(data.get("client_name")) or "Client Import",
                item_label=self._required_text(data, "item_label"),
                quantity=quantity,
                unit_price=unit_price,
                discount=discount,
                total_amount=total_amt,
                paid_amount=paid_amount,
                payment_method=self._optional_text(data.get("payment_method")) or "Espèces",
                payment_status=payment_status,
                source=RecordSource.EXCEL,
                status=RecordStatus.VALIDATED,
                created_by_user_id=user_id,
                updated_by_user_id=user_id,
            )
            
        if register_type in {"depenses", "expenses"}:
            amt = self._decimal(data.get("amount"), default=Decimal("0.00"))
            ref = str(data.get("reference") or f"IMP-DEPENSE-{uuid4().hex[:6].upper()}").strip()
            payment_status = self._payment_status(data.get("payment_status"))
            paid_amt = amt if payment_status == PaymentStatus.PAID else Decimal("0.00")
            
            return Expense(
                organization_id=organization_id,
                reference=ref,
                expense_date=self._date(data.get("expense_date")),
                document_type=ExpenseDocumentType.EXPENSE_RECEIPT,
                beneficiary=self._required_text(data, "beneficiary"),
                category=self._optional_text(data.get("category")) or "Divers",
                amount=amt,
                paid_amount=paid_amt,
                payment_method=self._optional_text(data.get("payment_method")) or "Espèces",
                payment_status=payment_status,
                source=RecordSource.EXCEL,
                status=RecordStatus.VALIDATED,
                created_by_user_id=user_id,
                updated_by_user_id=user_id,
            )
            
        if register_type in {"fournisseurs", "suppliers"}:
            return Supplier(
                organization_id=organization_id,
                name=self._required_text(data, "name"),
                category=self._optional_text(data.get("category")) or "Général",
                contact_name=self._optional_text(data.get("contact_name")),
                phone=self._optional_text(data.get("phone")),
                email=self._optional_text(data.get("email")),
                address=self._optional_text(data.get("address")),
                created_by_user_id=user_id,
                updated_by_user_id=user_id,
            )
            
        return Procedure(
            organization_id=organization_id,
            title=self._required_text(data, "title"),
            objective=self._optional_text(data.get("objective")),
            department=self._optional_text(data.get("department")),
            responsible_user_id=self._optional_text(data.get("responsible_user_id")),
            source=RecordSource.EXCEL,
            status=RecordStatus.VALIDATED,
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
        )

    async def rollback(self, session, organization_id: str, job_id: str) -> ImportJob:
        job = await session.scalar(
            select(ImportJob).where(
                ImportJob.id == job_id,
                ImportJob.organization_id == organization_id,
            )
        )
        if job is None:
            raise ApplicationError("import_not_found", "Import introuvable", 404)
        if job.status != ImportStatus.COMPLETED:
            raise ApplicationError(
                "rollback_unavailable",
                "Seul un import terminé peut être annulé",
                409,
            )
        model_map = {
            "offers": Offer,
            "products": Offer,
            "sales": Sale,
            "depenses": Expense,
            "expenses": Expense,
            "fournisseurs": Supplier,
            "suppliers": Supplier,
            "procedures": Procedure,
        }
        model = model_map.get(job.register_type, Offer)
        await session.execute(
            delete(model).where(
                model.organization_id == organization_id,
                model.id.in_(job.imported_record_ids),
            )
        )
        job.status = ImportStatus.ROLLED_BACK
        await session.commit()
        await session.refresh(job)
        return job

    @staticmethod
    def _required_text(data: dict[str, object], field: str) -> str:
        value = str(data.get(field) or "").strip()
        if not value:
            raise ValueError(f"{field} est requis")
        return value

    @staticmethod
    def _optional_text(value: object) -> str | None:
        text = str(value or "").strip()
        return text or None

    @staticmethod
    def _decimal(
        value: object,
        default: Decimal | None = None,
        nullable: bool = False,
    ) -> Decimal | None:
        if value in (None, ""):
            if nullable:
                return None
            if default is not None:
                return default
            raise ValueError("Valeur numérique requise")
        cleaned = str(value).replace(" ", "").replace("\u00a0", "").replace(",", ".")
        try:
            return Decimal(cleaned)
        except Exception:
            return default if default is not None else Decimal("0.00")

    @staticmethod
    def _date(value: object) -> date:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if not value or not str(value).strip():
            return date.today()
        raw = str(value).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                pass
        try:
            return date.fromisoformat(raw[:10])
        except Exception:
            return date.today()

    @staticmethod
    def _payment_status(value: object) -> PaymentStatus:
        normalized = str(value or "paid").strip().lower()
        if any(w in normalized for w in ["impayé", "impaye", "non payé", "non paye", "unpaid", "attente", "credit", "crédit"]):
            return PaymentStatus.UNPAID
        if any(w in normalized for w in ["partiel", "partial", "acompte"]):
            return PaymentStatus.PARTIAL
        return PaymentStatus.PAID
